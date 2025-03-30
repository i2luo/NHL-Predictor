# import requests
# from bs4 import BeautifulSoup
# import pandas as pd

# url = "https://www.hockey-reference.com/teams/EDM/2025_gamelog.html"
# response = requests.get(url)
# soup = BeautifulSoup(response.content, 'html.parser')

# # Find the table
# table = soup.find(name='table')

# # Use pandas to read the HTML table
# df = pd.read_html(str(table))[0]

# # Flatten multi-index columns (if applicable)
# df.columns = [' '.join(col).strip() if isinstance(col, tuple) else col for col in df.columns]

# # Save to Excel
# df.to_excel("oilers_gamelog.xlsx", index=False)

# print("Excel file created successfully!")

# import requests
# from bs4 import BeautifulSoup
# import pandas as pd

# #this is the url
# url = "https://www.hockey-reference.com/teams/EDM/2025_gamelog.html"
# response = requests.get(url)
# soup = BeautifulSoup(response.content, 'html.parser')

# #Find the table
# table = soup.find(name='table')

# #Remove all rows with class "thead"
# if table:
#     thead_rows = table.find_all('tr', class_='thead')
#     for row in thead_rows:
#         row.decompose()  # Remove the row from the DOM

# #Convert the modified table to string for pandas
# table_html = str(table) if table else ""

# #Use pandas to read the HTML table
# df = pd.read_html(table_html)[0]

# #Flatten multi-index columns (if applicable)
# df.columns = [' '.join(col).strip() if isinstance(col, tuple) else col for col in df.columns]

# #Save to Excel
# df.to_excel("oilers_gamelog.xlsx", index=False)

# print("Excel file created successfully!")

# import requests
# from bs4 import BeautifulSoup
# import pandas as pd

# #this is the url
# url = "https://www.hockey-reference.com/teams/EDM/2025_gamelog.html"
# response = requests.get(url)
# soup = BeautifulSoup(response.content, 'html.parser')

# # Find the table
# table = soup.find(name='table')

# # Remove all rows with class "thead"
# if table:
#     thead_rows = table.find_all('tr', class_='thead')
#     for row in thead_rows:
#         row.decompose()  # Remove the row from the DOM

# # Convert the modified table to string for pandas
# table_html = str(table) if table else ""

# # Use pandas to read the HTML table
# df = pd.read_html(table_html)[0]

# # Flatten multi-index columns (if applicable)
# df.columns = [' '.join(col).strip() if isinstance(col, tuple) else col for col in df.columns]

# # Save to Excel
# df.to_excel("oilers_gamelog.xlsx", index=False)

# print("Excel file created successfully!")

# import requests
# from bs4 import BeautifulSoup
# import pandas as pd
# import openpyxl


# #this is the url
# url = "https://www.hockey-reference.com/teams/EDM/2025_gamelog.html"
# response = requests.get(url)
# soup = BeautifulSoup(response.content, 'html.parser')

# # Find the table
# table = soup.find(name='table')

# # Remove all rows with class "thead"
# if table:
#     thead_rows = table.find_all('tr', class_='thead')
#     if thead_rows:
#         for row in thead_rows:
#             row.decompose()  # Remove the row from the DOM

# # Convert the modified table to string for pandas
# table_html = str(table) if table else ""

# # Use pandas to read the HTML table
# df = pd.read_html(table_html)[0]

# # Flatten multi-index columns (if applicable)
# df.columns = [' '.join(col).strip() if isinstance(col, tuple) else col for col in df.columns]

# # Save to Excel
# df.to_excel("oilers_gamelog.xlsx", index=False)

# print("Excel file created successfully!")






# For one team
# import requests
# from bs4 import BeautifulSoup
# import pandas as pd
# import openpyxl
# from io import StringIO

# # Define the team name (this will be added to the new column)
# team_name = "Vancouver Canucks"  # Change this based on the team

# # URL of the game log
# url = "https://www.hockey-reference.com/teams/VAN/2025_gamelog.html"
# response = requests.get(url)
# soup = BeautifulSoup(response.content, 'html.parser')

# # Find the table
# table = soup.find(name='table')

# # Remove all rows with class "thead"
# if table:
#     for row in table.find_all('tr', class_='thead'):
#         row.decompose()  # Remove from DOM

# # Convert to string and use StringIO to avoid the FutureWarning
# table_html = StringIO(str(table))

# # Read the table into a DataFrame
# df = pd.read_html(table_html)[0]

# # Flatten multi-index columns (if applicable)
# df.columns = [' '.join(col).strip() if isinstance(col, tuple) else col for col in df.columns]

# # Add a new column specifying the team name
# df.insert(0, "Team", team_name)

# # Define the Excel file path
# file_path = "future_games.xlsx"
# sheet_name = "Sheet1"

# try:
#     # Load existing workbook
#     with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
#         df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=writer.sheets[sheet_name].max_row, header=False)

#     print(f"Data for {team_name} appended successfully to Sheet1!")
# except FileNotFoundError:
#     # If file does not exist, create a new one
#     df.to_excel(file_path, index=False, sheet_name=sheet_name)
#     print(f"New Excel file created successfully with data for {team_name} in Sheet1!")

# # Load the existing Excel file
# wb = openpyxl.load_workbook(file_path)
# ws = wb.active  # Get the active worksheet

# # Modify the headers (including the new "Team" column)
# header_values = [
#     "Team", "RK", "GTM", "Date", "", "Opponents", "Results", "GF", "GA", "OT", "TEAM SOG", "TEAM PIM", "TEAM PPG", "TEAM PPO", "TEAM SHG",
#     "OPPONENT SOG", "OPPONENT PIM", "OPPONENT PPG", "OPPONENT PPO", "OPPONENT SHG", "FOW", "FOL", "FO%",
#     "CF", "CA", "CF%", "FF", "FA", "FF%", "oZS%", "PDO"
# ]

# # Apply headers
# for col, header in enumerate(header_values, start=1):
#     ws.cell(row=1, column=col, value=header)

# # Save the modified file
# wb.save(file_path)

# print(f"Excel file updated successfully with Team column for {team_name}!")





# # For all teams all games
# import requests
# from bs4 import BeautifulSoup
# import pandas as pd
# import openpyxl
# from io import StringIO

# # List of all NHL team abbreviations
# team_abbreviations = [
#     "ANA", "ARI", "BOS", "BUF", "CGY", "CAR", "CHI", "COL", "CBJ", "DAL", "DET", "EDM", "FLA",
#     "LAK", "MIN", "MTL", "NSH", "NJD", "NYI", "NYR", "OTT", "PHI", "PIT", "SJS", "SEA", "STL",
#     "TBL", "TOR", "VAN", "VGK", "WSH", "WPG"
# ]

# # Define the Excel file path
# file_path = "nhl_gamelogs.xlsx"
# sheet_name = "Sheet1"

# # Column headers
# header_values = [
#     "Team", "RK", "GTM", "Date", "", "Opponents", "Results", "GF", "GA", "OT", "TEAM SOG", "TEAM PIM", "TEAM PPG", 
#     "TEAM PPO", "TEAM SHG", "OPPONENT SOG", "OPPONENT PIM", "OPPONENT PPG", "OPPONENT PPO", "OPPONENT SHG", 
#     "FOW", "FOL", "FO%", "CF", "CA", "CF%", "FF", "FA", "FF%", "oZS%", "PDO"
# ]

# # Check if the file exists, if not create a new one with headers
# try:
#     wb = openpyxl.load_workbook(file_path)
#     ws = wb.active
#     print("Existing Excel file loaded.")
# except FileNotFoundError:
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.title = sheet_name
#     # Write headers
#     for col, header in enumerate(header_values, start=1):
#         ws.cell(row=1, column=col, value=header)
#     wb.save(file_path)
#     print("New Excel file created with headers.")

# # Process each team
# for team in team_abbreviations:
#     team_name = team  # You can replace this with full team names if needed
#     url = f"https://www.hockey-reference.com/teams/{team}/2024_gamelog.html"

#     print(f"Fetching data for {team_name}...")

#     # Fetch and parse the page
#     response = requests.get(url)
#     soup = BeautifulSoup(response.content, 'html.parser')

#     # Find the game log table
#     table = soup.find(name='table')

#     if not table:
#         print(f"Warning: No data found for {team_name}. Skipping...")
#         continue

#     # Remove all rows with class "thead"
#     for row in table.find_all('tr', class_='thead'):
#         row.decompose()  # Remove from DOM

#     # Convert to string and use StringIO to avoid FutureWarning
#     table_html = StringIO(str(table))

#     # Read the table into a DataFrame
#     df = pd.read_html(table_html)[0]

#     # Flatten multi-index columns if applicable
#     df.columns = [' '.join(col).strip() if isinstance(col, tuple) else col for col in df.columns]

#     # Add a new column specifying the team name
#     df.insert(0, "Team", team_name)

#     # Append the data to the existing Excel file
#     with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
#         df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=writer.sheets[sheet_name].max_row, header=False)

#     print(f"Data for {team_name} appended successfully!")

# print("All team data has been processed and saved.")



# Only first 59 for one team
# import requests
# from bs4 import BeautifulSoup
# import pandas as pd
# import openpyxl
# from io import StringIO

# # Define the team name (this will be added to the new column)
# team_name = "Edmonton Oilers"
# team_abbr = "EDM"  # Team abbreviation for URL

# # URL of the game log
# url = f"https://www.hockey-reference.com/teams/{team_abbr}/2025_gamelog.html"
# response = requests.get(url)
# soup = BeautifulSoup(response.content, 'html.parser')

# # Find the table
# table = soup.find(name='table')

# # Remove all rows with class "thead"
# if table:
#     for row in table.find_all('tr', class_='thead'):
#         row.decompose()

# # Convert to string and use StringIO to avoid the FutureWarning
# table_html = StringIO(str(table))

# # Read the table into a DataFrame
# df = pd.read_html(table_html)[0]

# # Flatten multi-index columns (if applicable)
# df.columns = [' '.join(col).strip() if isinstance(col, tuple) else col for col in df.columns]

# # Add a new column specifying the team name
# df.insert(0, "Team", team_name)

# # Define the Excel file path
# file_path = "oilers_gamelog.xlsx"
# sheet_name = "Sheet1"

# # Save the full data first
# df.to_excel(file_path, index=False, sheet_name=sheet_name)

# # Load the saved Excel file and delete rows 59 and onward
# wb = openpyxl.load_workbook(file_path)
# ws = wb.active  # Get the active worksheet

# # Modify the headers (including the new "Team" column)
# header_values = [
#     "Team", "RK", "GTM", "Date", "", "Opponents", "Results", "GF", "GA", "OT", "TEAM SOG", "TEAM PIM", "TEAM PPG", "TEAM PPO", "TEAM SHG",
#     "OPPONENT SOG", "OPPONENT PIM", "OPPONENT PPG", "OPPONENT PPO", "OPPONENT SHG", "FOW", "FOL", "FO%",
#     "CF", "CA", "CF%", "FF", "FA", "FF%", "oZS%", "PDO"
# ]

# # Apply headers
# for col, header in enumerate(header_values, start=1):
#     ws.cell(row=1, column=col, value=header)


# # Delete rows 60 and beyond (headers are in row 1)
# for row in range(60, ws.max_row + 1):
#     ws.delete_rows(60)

# # Save the modified file
# wb.save(file_path)

# print(f"Excel file updated successfully with Team column for {team_name}, and only games up to the 59th are kept!")




# import requests
# from bs4 import BeautifulSoup
# import pandas as pd
# import openpyxl
# from io import StringIO

# # List of NHL teams with their abbreviations
# teams = {
#     "Anaheim Ducks": "ANA", "Arizona Coyotes": "ARI", "Boston Bruins": "BOS", "Buffalo Sabres": "BUF",
#     "Calgary Flames": "CGY", "Carolina Hurricanes": "CAR", "Chicago Blackhawks": "CHI", "Colorado Avalanche": "COL",
#     "Columbus Blue Jackets": "CBJ", "Dallas Stars": "DAL", "Detroit Red Wings": "DET", "Edmonton Oilers": "EDM",
#     "Florida Panthers": "FLA", "Los Angeles Kings": "LAK", "Minnesota Wild": "MIN", "Montreal Canadiens": "MTL",
#     "Nashville Predators": "NSH", "New Jersey Devils": "NJD", "New York Islanders": "NYI", "New York Rangers": "NYR",
#     "Ottawa Senators": "OTT", "Philadelphia Flyers": "PHI", "Pittsburgh Penguins": "PIT", "San Jose Sharks": "SJS",
#     "Seattle Kraken": "SEA", "St. Louis Blues": "STL", "Tampa Bay Lightning": "TBL", "Toronto Maple Leafs": "TOR",
#     "Vancouver Canucks": "VAN", "Vegas Golden Knights": "VGK", "Washington Capitals": "WSH", "Winnipeg Jets": "WPG"
# }

# # Define the Excel file path
# file_path = "nhl_gamelog.xlsx"
# sheet_name = "Sheet1"

# # Create a blank DataFrame to store all teams' data
# all_games = pd.DataFrame()

# for team_name, team_abbr in teams.items():
#     print(f"Scraping {team_name}...")

#     # Construct the URL
#     url = f"https://www.hockey-reference.com/teams/{team_abbr}/2025_gamelog.html"
#     response = requests.get(url)
#     soup = BeautifulSoup(response.content, 'html.parser')

#     # Find the table
#     table = soup.find(name='table')

#     if table:
#         # Remove all rows with class "thead"
#         for row in table.find_all('tr', class_='thead'):
#             row.decompose()

#         # Convert table to string and parse with pandas
#         table_html = StringIO(str(table))
#         df = pd.read_html(table_html)[0]

#         # Flatten multi-index columns if applicable
#         df.columns = [' '.join(col).strip() if isinstance(col, tuple) else col for col in df.columns]

#         # Add a new column specifying the team name
#         df.insert(0, "Team", team_name)

#         # Keep only the first 59 games
#         df = df.iloc[:59]

#         # Append to the all_games DataFrame
#         all_games = pd.concat([all_games, df], ignore_index=True)

# # Save all data to an Excel file
# all_games.to_excel(file_path, index=False, sheet_name=sheet_name)

# # Load the Excel file to format headers
# wb = openpyxl.load_workbook(file_path)
# ws = wb.active

# # Define proper headers
# header_values = [
#     "Team", "RK", "GTM", "Date", "", "Opponents", "Results", "GF", "GA", "OT", "TEAM SOG", "TEAM PIM", "TEAM PPG", 
#     "TEAM PPO", "TEAM SHG", "OPPONENT SOG", "OPPONENT PIM", "OPPONENT PPG", "OPPONENT PPO", "OPPONENT SHG", 
#     "FOW", "FOL", "FO%", "CF", "CA", "CF%", "FF", "FA", "FF%", "oZS%", "PDO"
# ]

# # Apply headers in the first row
# # for col, header in enumerate(header_values, start=1):
# #     ws.cell(row=1, column=col, value=header)

# # Save the formatted file
# wb.save(file_path)

# print(f"Excel file '{file_path}' successfully created with all 32 teams' first 59 games!")







import requests
from bs4 import BeautifulSoup
import pandas as pd
import openpyxl
from io import StringIO

# List of NHL teams with their abbreviations
teams = {
    "Anaheim Ducks": "ANA", "Arizona Coyotes": "ARI", "Boston Bruins": "BOS", "Buffalo Sabres": "BUF",
    "Calgary Flames": "CGY", "Carolina Hurricanes": "CAR", "Chicago Blackhawks": "CHI", "Colorado Avalanche": "COL",
    "Columbus Blue Jackets": "CBJ", "Dallas Stars": "DAL", "Detroit Red Wings": "DET", "Edmonton Oilers": "EDM",
    "Florida Panthers": "FLA", "Los Angeles Kings": "LAK", "Minnesota Wild": "MIN", "Montreal Canadiens": "MTL",
    "Nashville Predators": "NSH", "New Jersey Devils": "NJD", "New York Islanders": "NYI", "New York Rangers": "NYR",
    "Ottawa Senators": "OTT", "Philadelphia Flyers": "PHI", "Pittsburgh Penguins": "PIT", "San Jose Sharks": "SJS",
    "Seattle Kraken": "SEA", "St. Louis Blues": "STL", "Tampa Bay Lightning": "TBL", "Toronto Maple Leafs": "TOR",
    "Vancouver Canucks": "VAN", "Vegas Golden Knights": "VGK", "Washington Capitals": "WSH", "Winnipeg Jets": "WPG"
}

# Define the Excel file path
file_path = "nhl_gamelog.xlsx"
sheet_name = "Sheet1"

# Create a blank DataFrame to store all teams' data
all_games = pd.DataFrame()

for team_name, team_abbr in teams.items():
    print(f"Scraping {team_name}...")

    # Construct the URL
    url = f"https://www.hockey-reference.com/teams/{team_abbr}/2025_gamelog.html"
    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'html.parser')

    # Find the table
    table = soup.find(name='table')

    if table:
        # Remove all rows with class "thead"
        for row in table.find_all('tr', class_='thead'):
            row.decompose()

        # Convert table to string and parse with pandas
        table_html = StringIO(str(table))
        df = pd.read_html(table_html)[0]

        # Flatten multi-index columns if applicable
        df.columns = [' '.join(col).strip() if isinstance(col, tuple) else col for col in df.columns]

        # Find the column name that corresponds to "Score Rslt" (7th column)
        # Column names might vary based on how pandas reads the HTML table
        result_column = df.columns[6]  # Assuming it's the 7th column (index 6)
        
        # Filter to include only rows with a value in the result column
        # This keeps only games that have been played
        df = df[df[result_column].notna()]

        # Add a new column specifying the team name
        df.insert(0, "Team", team_name)

        # Append to the all_games DataFrame
        all_games = pd.concat([all_games, df], ignore_index=True)

# Save all data to an Excel file
all_games.to_excel(file_path, index=False, sheet_name=sheet_name)

# Load the Excel file to format headers
wb = openpyxl.load_workbook(file_path)
ws = wb.active

# Define proper headers
header_values = [
    "Team", "RK", "GTM", "Date", "Home", "Opponents", "Results", "GF", "GA", "OT", "TEAM SOG", "TEAM PIM", "TEAM PPG", 
    "TEAM PPO", "TEAM SHG", "OPPONENT SOG", "OPPONENT PIM", "OPPONENT PPG", "OPPONENT PPO", "OPPONENT SHG", 
    "FOW", "FOL", "FO%", "CF", "CA", "CF%", "FF", "FA", "FF%", "oZS%", "PDO"
]
# Apply headers in the first row
for col, header in enumerate(header_values, start=1):
    ws.cell(row=1, column=col, value=header)

# Save the formatted file
wb.save(file_path)

print(f"Excel file '{file_path}' successfully created with all completed games for all 32 teams!")








# import pandas as pd
# import numpy as np
# from sklearn.model_selection import train_test_split
# from sklearn.preprocessing import StandardScaler, OneHotEncoder
# from tensorflow import keras
# from tensorflow.keras import layers
# import matplotlib.pyplot as plt

# # Step 1: Load and preprocess the data
# def load_and_preprocess_data(excel_file_path):
#     # Load the data
#     print("Reading Excel file...")
#     df = pd.read_excel(excel_file_path)
    
#     # Display initial columns and data types for debugging
#     print("DataFrame columns:", df.columns.tolist())
#     print("DataFrame dtypes sample:", df.dtypes.head())
    
#     # Convert date column to datetime if it exists and is not already datetime
#     if 'Date' in df.columns:
#         if not pd.api.types.is_datetime64_dtype(df['Date']):
#             print("Converting Date to datetime...")
#             df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
#     else:
#         print("Warning: Date column not found. Using index as proxy for time.")
#         df['Date'] = pd.date_range(start='2023-01-01', periods=len(df))
    
#     # Extract scores from Results column - with robust error handling
#     if 'Results' in df.columns and df['Results'].dtype == 'object':
#         print("Processing Results column...")
#         # Check if Results column follows the expected format
#         if df['Results'].str.contains('-').any():
#             try:
#                 # Extract Win/Loss information if format is like "W 3-2"
#                 if df['Results'].str.contains(' ').any():
#                     df['Win_Loss'] = df['Results'].str.split(' ', n=1).str[0]
#                     df['Score'] = df['Results'].str.split(' ', n=1).str[1]
#                 else:
#                     df['Score'] = df['Results']
                
#                 # Extract score components
#                 df['Team_Score'] = df['Score'].str.split('-').str[0].astype(int)
#                 df['Opponent_Score'] = df['Score'].str.split('-').str[1].astype(int)
#             except Exception as e:
#                 print(f"Error parsing Results column: {e}")
#                 # Use GF and GA as fallback
#                 print("Using GF and GA columns as direct score indicators.")
#                 df['Team_Score'] = df['GF'] if 'GF' in df.columns else 0
#                 df['Opponent_Score'] = df['GA'] if 'GA' in df.columns else 0
#         else:
#             print("Results column doesn't contain expected format. Using GF and GA columns.")
#             df['Team_Score'] = df['GF'] if 'GF' in df.columns else 0
#             df['Opponent_Score'] = df['GA'] if 'GA' in df.columns else 0
#     else:
#         print("Results column not found or not string type. Using GF and GA columns directly.")
#         # Use GF and GA columns directly if they exist
#         df['Team_Score'] = df['GF'] if 'GF' in df.columns else 0
#         df['Opponent_Score'] = df['GA'] if 'GA' in df.columns else 0
    
#     # Create features for time trends (days since season start)
#     season_start = df['Date'].min()
#     df['Days_Since_Season_Start'] = (df['Date'] - season_start).dt.days
    
#     # Create team form features (rolling averages)
#     print("Creating team rolling statistics...")
    
#     # Ensure Team column exists
#     if 'Team' not in df.columns:
#         print("Warning: Team column not found. Using first column as team identifier.")
#         df['Team'] = df.iloc[:, 0]
    
#     teams = df['Team'].unique()
#     print(f"Found {len(teams)} unique teams.")
    
#     # Create empty dataframes to store team stats
#     team_rolling_stats = pd.DataFrame()
    
#     # Define columns that should be available for rolling stats
#     stat_columns = [
#         'GF', 'GA', 'TEAM SOG', 'OPPONENT SOG', 'TEAM PPG', 'TEAM PIM',
#         'CF%', 'FF%', 'FO%', 'PDO'
#     ]
    
#     # Check which columns actually exist in the dataframe
#     available_columns = [col for col in stat_columns if col in df.columns]
#     print(f"Available stat columns: {available_columns}")
    
#     if len(available_columns) < 3:
#         print("Warning: Few stat columns available. Will use basic scoring stats.")
        
#         # If key columns missing, create some basic ones from available data
#         if 'GF' not in df.columns and 'Team_Score' in df.columns:
#             df['GF'] = df['Team_Score']
#         if 'GA' not in df.columns and 'Opponent_Score' in df.columns:
#             df['GA'] = df['Opponent_Score']
        
#         # Update available columns list
#         available_columns = [col for col in stat_columns if col in df.columns]
    
#     for team in teams:
#         team_data = df[df['Team'] == team].sort_values('Date')
        
#         if len(team_data) < 2:
#             print(f"Warning: Team {team} has very few data points. Skipping rolling statistics.")
#             continue
        
#         # Calculate rolling averages for available stats (last 5 games)
#         window = min(5, len(team_data) - 1)  # Ensure window isn't larger than data
        
#         # Create a dictionary for rolling stats
#         rolling_dict = {
#             'Team': team,
#             'Date': team_data['Date']
#         }
        
#         # Add rolling statistics for each available column
#         for col in available_columns:
#             if col in team_data.columns and pd.api.types.is_numeric_dtype(team_data[col]):
#                 col_name = f"Rolling_{col.replace('%', 'Pct').replace(' ', '_')}"
#                 rolling_dict[col_name] = team_data[col].rolling(window, min_periods=1).mean()
        
#         # Create DataFrame from the dictionary
#         rolling_stats = pd.DataFrame(rolling_dict)
        
#         # Concatenate to the main rolling stats DataFrame
#         team_rolling_stats = pd.concat([team_rolling_stats, rolling_stats], ignore_index=True)
    
#     # Check if we have rolling stats before merging
#     if len(team_rolling_stats) > 0:
#         print(f"Merging {len(team_rolling_stats)} rows of rolling statistics.")
#         # Merge rolling stats back to the main dataframe
#         df = pd.merge(df, team_rolling_stats, on=['Team', 'Date'], how='left')
#     else:
#         print("Warning: No rolling statistics were generated. Using static features only.")
    
#     # Fill NaN values
#     print("Handling missing values...")
#     for col in df.select_dtypes(include=[np.number]).columns:
#         if df[col].isna().any():
#             df[col] = df[col].fillna(df[col].mean())
    
#     return df

# # Step 2: Feature engineering
# def engineer_features(df):
#     print("Engineering features...")
    
#     # Identify the rolling stats columns that were created
#     rolling_cols = [col for col in df.columns if col.startswith('Rolling_')]
#     print(f"Available rolling stat features: {rolling_cols}")
    
#     # If we have at least 3 rolling features, use them; otherwise, use basic features
#     if len(rolling_cols) >= 3:
#         features = rolling_cols
#     else:
#         # Fallback to basic features
#         print("Using basic features due to limited rolling statistics")
#         numeric_cols = df.select_dtypes(include=[np.number]).columns.tolist()
#         # Exclude target variables and unnecessary columns
#         features = [col for col in numeric_cols if col not in ['Team_Score', 'Opponent_Score', 'index']]
    
#     # Always include time feature if available
#     if 'Days_Since_Season_Start' in df.columns:
#         if 'Days_Since_Season_Start' not in features:
#             features.append('Days_Since_Season_Start')
    
#     print(f"Selected {len(features)} features: {features[:min(5, len(features))]}...")
    
#     # Add team encoding
#     print("Encoding team names...")
#     encoder = OneHotEncoder(sparse_output=False, drop='first')
    
#     # Reshape to ensure 2D array for encoding
#     team_data = df[['Team']].values
#     team_encoded = encoder.fit_transform(team_data)
#     team_encoded_df = pd.DataFrame(
#         team_encoded, 
#         columns=[f'Team_{i}' for i in range(team_encoded.shape[1])]
#     )
    
#     # Prepare feature matrix
#     X = pd.concat([df[features].reset_index(drop=True), team_encoded_df.reset_index(drop=True)], axis=1)
    
#     # Target variables: Team score and opponent score
#     y_team = df['Team_Score']
#     y_opponent = df['Opponent_Score']
    
#     return X, y_team, y_opponent, encoder, features

# # Step 3: Build and train the neural network model
# def build_and_train_model(X, y_team, y_opponent):
#     print("Preparing data for model training...")
#     # Split the data into training and testing sets
#     X_train, X_test, y_team_train, y_team_test, y_opponent_train, y_opponent_test = train_test_split(
#         X, y_team, y_opponent, test_size=0.2, random_state=42
#     )
    
#     # Scale the features
#     scaler = StandardScaler()
#     X_train_scaled = scaler.fit_transform(X_train)
#     X_test_scaled = scaler.transform(X_test)
    
#     # Convert target variables to numpy arrays
#     y_team_train = np.array(y_team_train)
#     y_team_test = np.array(y_team_test)
#     y_opponent_train = np.array(y_opponent_train)
#     y_opponent_test = np.array(y_opponent_test)
    
#     print(f"Training data shape: {X_train_scaled.shape}")
#     print(f"Feature count: {X_train_scaled.shape[1]}")
    
#     # Build the model
#     print("Building neural network model...")
#     inputs = keras.Input(shape=(X_train_scaled.shape[1],))
    
#     # Start with more nodes and gradually reduce
#     x = layers.Dense(128, activation='relu')(inputs)
#     x = layers.BatchNormalization()(x)
#     x = layers.Dropout(0.3)(x)
    
#     x = layers.Dense(64, activation='relu')(x)
#     x = layers.BatchNormalization()(x)
#     x = layers.Dropout(0.3)(x)
    
#     x = layers.Dense(32, activation='relu')(x)
#     x = layers.Dropout(0.2)(x)
    
#     # Two output heads for team score and opponent score
#     team_score_output = layers.Dense(1, name='team_score')(x)
#     opponent_score_output = layers.Dense(1, name='opponent_score')(x)
    
#     model = keras.Model(inputs=inputs, outputs=[team_score_output, opponent_score_output])
    
#     # Compile the model
#     model.compile(
#         optimizer=keras.optimizers.Adam(learning_rate=0.001),
#         loss={'team_score': 'mse', 'opponent_score': 'mse'},
#         metrics={'team_score': 'mae', 'opponent_score': 'mae'}
#     )
    
#     # Define callbacks
#     callbacks = [
#         keras.callbacks.EarlyStopping(
#             patience=15, 
#             restore_best_weights=True,
#             monitor='val_loss'
#         ),
#         keras.callbacks.ReduceLROnPlateau(
#             factor=0.5, 
#             patience=5,
#             min_lr=0.00001
#         )
#     ]
    
#     # Train the model
#     print("Training the model (this may take a while)...")
#     history = model.fit(
#         X_train_scaled,
#         {'team_score': y_team_train, 'opponent_score': y_opponent_train},
#         epochs=100,
#         batch_size=32,
#         validation_split=0.2,
#         callbacks=callbacks,
#         verbose=1
#     )
    
#     # Evaluate the model
#     print("Evaluating model performance...")
#     eval_results = model.evaluate(
#         X_test_scaled,
#         {'team_score': y_team_test, 'opponent_score': y_opponent_test},
#         verbose=0
#     )
    
#     print(f"Test Loss (Team Score): {eval_results[1]:.4f}")
#     print(f"Test MAE (Team Score): {eval_results[3]:.4f}")
#     print(f"Test Loss (Opponent Score): {eval_results[2]:.4f}")
#     print(f"Test MAE (Opponent Score): {eval_results[4]:.4f}")
    
#     # Plot training history
#     print("Generating training history plots...")
#     plt.figure(figsize=(12, 8))
    
#     plt.subplot(2, 2, 1)
#     plt.plot(history.history['team_score_loss'])
#     plt.plot(history.history['val_team_score_loss'])
#     plt.title('Team Score Loss')
#     plt.ylabel('Loss (MSE)')
#     plt.xlabel('Epoch')
#     plt.legend(['Train', 'Validation'], loc='upper right')
    
#     plt.subplot(2, 2, 2)
#     plt.plot(history.history['team_score_mae'])
#     plt.plot(history.history['val_team_score_mae'])
#     plt.title('Team Score MAE')
#     plt.ylabel('MAE')
#     plt.xlabel('Epoch')
#     plt.legend(['Train', 'Validation'], loc='upper right')
    
#     plt.subplot(2, 2, 3)
#     plt.plot(history.history['opponent_score_loss'])
#     plt.plot(history.history['val_opponent_score_loss'])
#     plt.title('Opponent Score Loss')
#     plt.ylabel('Loss (MSE)')
#     plt.xlabel('Epoch')
#     plt.legend(['Train', 'Validation'], loc='upper right')
    
#     plt.subplot(2, 2, 4)
#     plt.plot(history.history['opponent_score_mae'])
#     plt.plot(history.history['val_opponent_score_mae'])
#     plt.title('Opponent Score MAE')
#     plt.ylabel('MAE')
#     plt.xlabel('Epoch')
#     plt.legend(['Train', 'Validation'], loc='upper right')
    
#     plt.tight_layout()
#     plt.savefig('training_history.png')
#     plt.show()
    
#     return model, scaler, history

# # Step 4: Make predictions for future games
# def predict_game_score(model, scaler, encoder, features, team, opponent_stats, days_since_season_start):
#     print(f"Generating prediction for {team}...")
#     # Create a dataframe for the prediction
#     pred_data = {
#         'Team': [team],
#         'Days_Since_Season_Start': [days_since_season_start]
#     }
    
#     # Add stats from the opponent_stats dictionary
#     for feature in features:
#         if feature in opponent_stats:
#             pred_data[feature] = [opponent_stats[feature]]
#         else:
#             # If a feature is missing, add a default value
#             print(f"Warning: Feature {feature} not provided. Using default value.")
#             pred_data[feature] = [0.0]
    
#     team_df = pd.DataFrame(pred_data)
    
#     # Encode the team
#     team_encoded = encoder.transform(team_df[['Team']])
#     team_encoded_df = pd.DataFrame(
#         team_encoded, 
#         columns=[f'Team_{i}' for i in range(team_encoded.shape[1])]
#     )
    
#     # Get only the features that were used during training
#     X_pred_features = pd.DataFrame()
#     for feature in features:
#         if feature in team_df.columns:
#             X_pred_features[feature] = team_df[feature]
    
#     # Combine features
#     X_pred = pd.concat([X_pred_features.reset_index(drop=True), team_encoded_df.reset_index(drop=True)], axis=1)
    
#     # Scale features
#     X_pred_scaled = scaler.transform(X_pred)
    
#     # Make prediction
#     predictions = model.predict(X_pred_scaled)
    
#     # Round to nearest integer and ensure non-negative
#     team_score = max(0, round(float(predictions[0][0])))
#     opponent_score = max(0, round(float(predictions[1][0])))
    
#     return team_score, opponent_score

# # Step 5: Model evaluation with visualization
# def evaluate_model_with_viz(model, X_test_scaled, y_team_test, y_opponent_test):
#     print("Generating prediction visualizations...")
#     # Make predictions on test data
#     predictions = model.predict(X_test_scaled)
#     team_preds = predictions[0].flatten()
#     opponent_preds = predictions[1].flatten()
    
#     # Create a figure for visualization
#     plt.figure(figsize=(15, 10))
    
#     # Scatter plot of actual vs predicted team scores
#     plt.subplot(2, 2, 1)
#     plt.scatter(y_team_test, team_preds, alpha=0.5)
#     plt.plot([0, max(y_team_test)], [0, max(y_team_test)], 'r--')
#     plt.xlabel('Actual Team Score')
#     plt.ylabel('Predicted Team Score')
#     plt.title('Team Score: Actual vs Predicted')
    
#     # Scatter plot of actual vs predicted opponent scores
#     plt.subplot(2, 2, 2)
#     plt.scatter(y_opponent_test, opponent_preds, alpha=0.5)
#     plt.plot([0, max(y_opponent_test)], [0, max(y_opponent_test)], 'r--')
#     plt.xlabel('Actual Opponent Score')
#     plt.ylabel('Predicted Opponent Score')
#     plt.title('Opponent Score: Actual vs Predicted')
    
#     # Histogram of residuals for team scores
#     plt.subplot(2, 2, 3)
#     residuals_team = y_team_test - team_preds
#     plt.hist(residuals_team, bins=20)
#     plt.xlabel('Residual')
#     plt.ylabel('Frequency')
#     plt.title('Residuals for Team Score')
    
#     # Histogram of residuals for opponent scores
#     plt.subplot(2, 2, 4)
#     residuals_opponent = y_opponent_test - opponent_preds
#     plt.hist(residuals_opponent, bins=20)
#     plt.xlabel('Residual')
#     plt.ylabel('Frequency')
#     plt.title('Residuals for Opponent Score')
    
#     plt.tight_layout()
#     plt.savefig('model_evaluation.png')
#     plt.show()
    
#     # Calculate and print evaluation metrics
#     team_mae = np.mean(np.abs(residuals_team))
#     opponent_mae = np.mean(np.abs(residuals_opponent))
#     team_mse = np.mean(residuals_team ** 2)
#     opponent_mse = np.mean(residuals_opponent ** 2)
    
#     print(f"Team Score - MAE: {team_mae:.4f}, MSE: {team_mse:.4f}")
#     print(f"Opponent Score - MAE: {opponent_mae:.4f}, MSE: {opponent_mse:.4f}")
    
#     # Calculate accuracy for predicting win/loss/tie
#     actual_result = np.where(y_team_test > y_opponent_test, 'W', 
#                              np.where(y_team_test < y_opponent_test, 'L', 'T'))
#     pred_result = np.where(team_preds > opponent_preds, 'W', 
#                            np.where(team_preds < opponent_preds, 'L', 'T'))
#     result_accuracy = np.mean(actual_result == pred_result)
    
#     print(f"Win/Loss/Tie Prediction Accuracy: {result_accuracy:.4f}")
    
#     return team_mae, opponent_mae, result_accuracy

# # Main execution flow
# def main():
#     # File path to your NHL data
#     excel_file_path = 'nhl_gamelogs.xlsx'
    
#     try:
#         # Load and preprocess the data
#         print("Starting data loading and preprocessing...")
#         df = load_and_preprocess_data(excel_file_path)
        
#         # Check if we have enough data after preprocessing
#         if len(df) < 100:
#             print(f"Warning: Dataset has only {len(df)} rows after preprocessing. Model may not be accurate.")
        
#         # Engineer features
#         print("Starting feature engineering...")
#         X, y_team, y_opponent, encoder, features = engineer_features(df)
        
#         # Build and train the model
#         print("Starting model training...")
#         model, scaler, history = build_and_train_model(X, y_team, y_opponent)
        
#         # Split data for evaluation
#         _, X_test, _, y_team_test, _, y_opponent_test = train_test_split(
#             X, y_team, y_opponent, test_size=0.2, random_state=42
#         )
#         X_test_scaled = scaler.transform(X_test)
        
#         # Evaluate model with visualizations
#         evaluate_model_with_viz(model, X_test_scaled, y_team_test, y_opponent_test)
        
#         # Example prediction for a future game
#         print("\nGenerating example prediction...")
#         team = df['Team'].iloc[0]  # Use first team from data as example
        
#         # Create example stats based on averages from the data
#         example_stats = {}
#         for feature in features:
#             if feature in df.columns and feature != 'Days_Since_Season_Start':
#                 example_stats[feature] = df[feature].mean()
        
#         # Set days since season start to an appropriate value
#         days_since_season_start = 180  # Example: mid-season
        
#         # Make prediction
#         team_score, opponent_score = predict_game_score(
#             model, scaler, encoder, features, team, example_stats, days_since_season_start
#         )
        
#         print(f"\nPredicted score for {team}: {team_score}-{opponent_score}")
        
#         # Save the model and related components
#         print("Saving model and components...")
#         model.save('nhl_score_prediction_model.h5')
        
#         # Save feature names and encoder for future use
#         import pickle
#         with open('model_scaler.pkl', 'wb') as f:
#             pickle.dump(scaler, f)
#         with open('model_encoder.pkl', 'wb') as f:
#             pickle.dump(encoder, f)
#         with open('model_features.pkl', 'wb') as f:
#             pickle.dump(features, f)
            
#         print("Model and components saved successfully.")
#         print("\nTo make predictions with this model, use the predict_game_score function.")
        
#     except Exception as e:
#         print(f"An error occurred: {e}")
#         import traceback
#         traceback.print_exc()

# if __name__ == "__main__":
#     main()













# import pandas as pd
# import numpy as np
# import datetime
# import os

# # Dictionary to map team abbreviations to full names
# TEAM_ABBR_TO_NAME = {
#     'VAN': 'Vancouver Canucks',
#     'EDM': 'Edmonton Oilers',
#     'CGY': 'Calgary Flames',
#     'SEA': 'Seattle Kraken',
#     'VGK': 'Vegas Golden Knights',
#     'LAK': 'Los Angeles Kings',
#     'SJS': 'San Jose Sharks',
#     'ANA': 'Anaheim Ducks',
#     'COL': 'Colorado Avalanche',
#     'MIN': 'Minnesota Wild',
#     'DAL': 'Dallas Stars',
#     'WPG': 'Winnipeg Jets',
#     'STL': 'St. Louis Blues',
#     'NSH': 'Nashville Predators',
#     'CHI': 'Chicago Blackhawks',
#     'ARI': 'Arizona Coyotes',
#     'TOR': 'Toronto Maple Leafs',
#     'BOS': 'Boston Bruins',
#     'TBL': 'Tampa Bay Lightning',
#     'FLA': 'Florida Panthers',
#     'DET': 'Detroit Red Wings',
#     'OTT': 'Ottawa Senators',
#     'MTL': 'Montreal Canadiens',
#     'BUF': 'Buffalo Sabres',
#     'NYR': 'New York Rangers',
#     'NYI': 'New York Islanders',
#     'NJD': 'New Jersey Devils',
#     'PIT': 'Pittsburgh Penguins',
#     'PHI': 'Philadelphia Flyers',
#     'WSH': 'Washington Capitals',
#     'CAR': 'Carolina Hurricanes',
#     'CBJ': 'Columbus Blue Jackets',
# }

# # Function to load the schedule data from Excel
# def load_schedule(file_path):
#     """
#     Load the schedule data from Excel file
#     Column 1: Date
#     Column 2: Home Team
#     Column 3: Away Team
#     """
#     try:
#         schedule_df = pd.read_excel(file_path)
        
#         # Rename columns if they don't have proper names
#         if schedule_df.columns[0] == 0 or 'Unnamed' in str(schedule_df.columns[0]):
#             schedule_df.columns = ['Date', 'Home_Team', 'Away_Team']
        
#         # Convert date strings to datetime objects if needed
#         if not pd.api.types.is_datetime64_any_dtype(schedule_df['Date']):
#             schedule_df['Date'] = pd.to_datetime(schedule_df['Date'])
        
#         # Print first few rows for debugging
#         print("Schedule data (first 5 rows):")
#         print(schedule_df.head())
        
#         # Add a column to identify if Canucks are home or away
#         schedule_df['Canucks_Home'] = schedule_df.apply(
#             lambda row: 'VAN' in str(row['Home_Team']) or 'Vancouver' in str(row['Home_Team']), 
#             axis=1
#         )
        
#         # Convert team abbreviations to full names
#         schedule_df['Home_Team_Full'] = schedule_df['Home_Team'].apply(
#             lambda x: TEAM_ABBR_TO_NAME.get(x, x) if isinstance(x, str) else x
#         )
        
#         schedule_df['Away_Team_Full'] = schedule_df['Away_Team'].apply(
#             lambda x: TEAM_ABBR_TO_NAME.get(x, x) if isinstance(x, str) else x
#         )
        
#         return schedule_df
    
#     except Exception as e:
#         print(f"Error loading schedule: {e}")
#         print("File structure:")
#         try:
#             temp_df = pd.read_excel(file_path)
#             print(temp_df.head())
#             print(f"Columns: {temp_df.columns.tolist()}")
#         except:
#             print("Unable to read file structure")
#         raise

# # Function to load team statistics
# def load_team_stats():
#     """
#     In a real implementation, this would load statistics from an API or database.
#     For this example, we'll create synthetic data for NHL teams.
#     """
#     # NHL teams
#     teams = [
#         'Vancouver Canucks', 'Edmonton Oilers', 'Calgary Flames', 'Seattle Kraken',
#         'Vegas Golden Knights', 'Los Angeles Kings', 'San Jose Sharks', 'Anaheim Ducks',
#         'Colorado Avalanche', 'Minnesota Wild', 'Dallas Stars', 'Winnipeg Jets',
#         'St. Louis Blues', 'Nashville Predators', 'Chicago Blackhawks', 'Arizona Coyotes',
#         'Toronto Maple Leafs', 'Boston Bruins', 'Tampa Bay Lightning', 'Florida Panthers',
#         'Detroit Red Wings', 'Ottawa Senators', 'Montreal Canadiens', 'Buffalo Sabres',
#         'New York Rangers', 'New York Islanders', 'New Jersey Devils', 'Pittsburgh Penguins',
#         'Philadelphia Flyers', 'Washington Capitals', 'Carolina Hurricanes', 'Columbus Blue Jackets'
#     ]
    
#     # Create a dictionary for team statistics
#     team_stats = {}
    
#     # Set seed for reproducibility
#     np.random.seed(42)
    
#     # Define mean values for different team tiers
#     top_tier = {'GF': 3.4, 'GA': 2.6, 'SOG': 33, 'SOG_against': 28, 'PPG': 1.1, 'PIM': 8, 'CF': 53, 'FF': 52, 'FO': 52, 'PDO': 101, 'OT_WIN_PCT': 0.62}
#     mid_tier = {'GF': 3.0, 'GA': 3.0, 'SOG': 30, 'SOG_against': 30, 'PPG': 0.8, 'PIM': 9, 'CF': 50, 'FF': 50, 'FO': 50, 'PDO': 100, 'OT_WIN_PCT': 0.5}
#     low_tier = {'GF': 2.6, 'GA': 3.4, 'SOG': 28, 'SOG_against': 33, 'PPG': 0.5, 'PIM': 10, 'CF': 47, 'FF': 48, 'FO': 48, 'PDO': 99, 'OT_WIN_PCT': 0.38}
    
#     # Define team tiers for 2025 season (with Canucks as a good team for this example)
#     top_teams = ['Vancouver Canucks', 'Edmonton Oilers', 'Colorado Avalanche', 'Vegas Golden Knights', 
#                 'Toronto Maple Leafs', 'Boston Bruins', 'Tampa Bay Lightning', 'Carolina Hurricanes', 
#                 'New York Rangers', 'Dallas Stars', 'Florida Panthers']
    
#     bottom_teams = ['Chicago Blackhawks', 'San Jose Sharks', 'Anaheim Ducks', 'Montreal Canadiens', 
#                    'Philadelphia Flyers', 'Columbus Blue Jackets', 'Buffalo Sabres']
    
#     # Generate stats for each team
#     for team in teams:
#         if team in top_teams:
#             base = top_tier
#             tier_factor = 1.2
#         elif team in bottom_teams:
#             base = low_tier
#             tier_factor = 0.8
#         else:
#             base = mid_tier
#             tier_factor = 1.0
        
#         # Add some random variation to make teams unique
#         team_stats[team] = {
#             'Rolling_GF': base['GF'] * (0.9 + 0.2 * np.random.random()),
#             'Rolling_GA': base['GA'] * (0.9 + 0.2 * np.random.random()),
#             'Rolling_TEAM_SOG': base['SOG'] * (0.9 + 0.2 * np.random.random()),
#             'Rolling_OPPONENT_SOG': base['SOG_against'] * (0.9 + 0.2 * np.random.random()),
#             'Rolling_TEAM_PPG': base['PPG'] * (0.8 + 0.4 * np.random.random()),
#             'Rolling_TEAM_PIM': base['PIM'] * (0.8 + 0.4 * np.random.random()),
#             'Rolling_CFPct': base['CF'] * (0.95 + 0.1 * np.random.random()),
#             'Rolling_FFPct': base['FF'] * (0.95 + 0.1 * np.random.random()),
#             'Rolling_FOPct': base['FO'] * (0.95 + 0.1 * np.random.random()),
#             'Rolling_PDO': base['PDO'] * (0.99 + 0.02 * np.random.random()),
#             'Team_Strength': tier_factor * (0.9 + 0.2 * np.random.random()),
#             'OT_Win_Pct': base['OT_WIN_PCT'] * (0.9 + 0.2 * np.random.random())  # Added OT win percentage
#         }
    
#     # Add abbreviations as keys too
#     abbr_stats = {}
#     for abbr, name in TEAM_ABBR_TO_NAME.items():
#         if name in team_stats:
#             abbr_stats[abbr] = team_stats[name]
    
#     # Merge dictionaries
#     team_stats.update(abbr_stats)
    
#     return team_stats

# # Function to create feature vector for prediction
# def create_feature_vector(home_team, away_team, team_stats, days_since_season_start):
#     """
#     Create the feature vector for the prediction model
#     """
#     # Try to get team stats, handling both full names and abbreviations
#     try:
#         home_stats = team_stats[home_team]
#     except KeyError:
#         # Try to convert from abbreviation to full name
#         home_full = TEAM_ABBR_TO_NAME.get(home_team, home_team)
#         try:
#             home_stats = team_stats[home_full]
#         except KeyError:
#             print(f"Warning: Could not find stats for team: {home_team} or {home_full}")
#             print(f"Available teams: {list(team_stats.keys())[:5]}...")
#             # Use default stats
#             home_stats = {
#                 'Team_Strength': 1.0,
#                 'Rolling_GF': 3.0,
#                 'Rolling_GA': 3.0,
#                 'Rolling_TEAM_SOG': 30,
#                 'Rolling_OPPONENT_SOG': 30,
#                 'Rolling_TEAM_PPG': 0.8,
#                 'Rolling_CFPct': 50,
#                 'Rolling_PDO': 100,
#                 'OT_Win_Pct': 0.5
#             }
    
#     try:
#         away_stats = team_stats[away_team]
#     except KeyError:
#         # Try to convert from abbreviation to full name
#         away_full = TEAM_ABBR_TO_NAME.get(away_team, away_team)
#         try:
#             away_stats = team_stats[away_full]
#         except KeyError:
#             print(f"Warning: Could not find stats for team: {away_team} or {away_full}")
#             print(f"Available teams: {list(team_stats.keys())[:5]}...")
#             # Use default stats
#             away_stats = {
#                 'Team_Strength': 1.0,
#                 'Rolling_GF': 3.0,
#                 'Rolling_GA': 3.0,
#                 'Rolling_TEAM_SOG': 30,
#                 'Rolling_OPPONENT_SOG': 30,
#                 'Rolling_TEAM_PPG': 0.8,
#                 'Rolling_CFPct': 50,
#                 'Rolling_PDO': 100,
#                 'OT_Win_Pct': 0.5
#             }
    
#     # Create feature vector
#     features = {
#         'Days_Since_Season_Start': days_since_season_start,
#         'Home_Team_Strength': home_stats['Team_Strength'],
#         'Away_Team_Strength': away_stats['Team_Strength'],
#         'Home_Rolling_GF': home_stats['Rolling_GF'],
#         'Home_Rolling_GA': home_stats['Rolling_GA'],
#         'Away_Rolling_GF': away_stats['Rolling_GF'],
#         'Away_Rolling_GA': away_stats['Rolling_GA'],
#         'Home_Rolling_SOG': home_stats['Rolling_TEAM_SOG'],
#         'Home_Rolling_SOG_Against': home_stats['Rolling_OPPONENT_SOG'],
#         'Away_Rolling_SOG': away_stats['Rolling_TEAM_SOG'],
#         'Away_Rolling_SOG_Against': away_stats['Rolling_OPPONENT_SOG'],
#         'Home_Rolling_PPG': home_stats['Rolling_TEAM_PPG'],
#         'Away_Rolling_PPG': away_stats['Rolling_TEAM_PPG'],
#         'Home_Rolling_CF': home_stats['Rolling_CFPct'],
#         'Away_Rolling_CF': away_stats['Rolling_CFPct'],
#         'Home_Rolling_PDO': home_stats['Rolling_PDO'],
#         'Away_Rolling_PDO': away_stats['Rolling_PDO'],
#         'Home_OT_Win_Pct': home_stats.get('OT_Win_Pct', 0.5),
#         'Away_OT_Win_Pct': away_stats.get('OT_Win_Pct', 0.5)
#     }
    
#     return features

# # Predict score using Poisson distribution with team statistics and resolve ties
# def predict_score(features):
#     """
#     Predict scores based on team stats using a modified Poisson model
#     Includes overtime resolution for tied games
#     """
#     # Base expected goals (league average is around 3.0 per team)
#     base_home_expected = 3.1  # Home teams score slightly more on average
#     base_away_expected = 2.9
    
#     # Apply modifiers based on team strengths and stats
#     home_expected = base_home_expected * (
#         features['Home_Team_Strength'] * 0.5 + 
#         (features['Home_Rolling_GF'] / features['Away_Rolling_GA']) * 0.3 +
#         (features['Home_Rolling_SOG'] / features['Away_Rolling_SOG_Against']) * 0.1 +
#         (features['Home_Rolling_CF'] / 50) * 0.05 +
#         (features['Home_Rolling_PDO'] / 100) * 0.05
#     )
    
#     away_expected = base_away_expected * (
#         features['Away_Team_Strength'] * 0.5 + 
#         (features['Away_Rolling_GF'] / features['Home_Rolling_GA']) * 0.3 +
#         (features['Away_Rolling_SOG'] / features['Home_Rolling_SOG_Against']) * 0.1 +
#         (features['Away_Rolling_CF'] / 50) * 0.05 +
#         (features['Away_Rolling_PDO'] / 100) * 0.05
#     )
    
#     # Add small random variation
#     home_expected *= (0.9 + 0.2 * np.random.random())
#     away_expected *= (0.9 + 0.2 * np.random.random())
    
#     # Generate scores using Poisson distribution
#     home_score = np.random.poisson(home_expected)
#     away_score = np.random.poisson(away_expected)
    
#     # Check if the game is tied after regulation
#     is_overtime = False
#     if home_score == away_score:
#         is_overtime = True
#         # Determine overtime winner based on OT win percentages
#         home_ot_win_pct = features['Home_OT_Win_Pct']
#         away_ot_win_pct = features['Away_OT_Win_Pct']
        
#         # Normalize to make sure probabilities sum to 1
#         total = home_ot_win_pct + away_ot_win_pct
#         home_ot_win_pct /= total
#         away_ot_win_pct /= total
        
#         # Determine winner
#         if np.random.random() < home_ot_win_pct:
#             home_score += 1  # Home team wins in OT
#         else:
#             away_score += 1  # Away team wins in OT
    
#     return home_score, away_score, is_overtime

# # Main function to predict all games
# def predict_canucks_games(schedule_file):
#     """
#     Predict all Vancouver Canucks games in the schedule
#     """
#     # Load the schedule
#     schedule = load_schedule(schedule_file)
    
#     # Load team statistics
#     team_stats = load_team_stats()
    
#     # Get current date
#     current_date = datetime.datetime.now()
    
#     # Assume season started on October 1, 2024
#     season_start = datetime.datetime(2024, 10, 1)
    
#     # Create results dataframe
#     results = []
    
#     # Process each game
#     for idx, game in schedule.iterrows():
#         # Skip games that have already happened
#         if game['Date'] < current_date:
#             continue
        
#         # Calculate days since season start
#         days_since_start = (game['Date'] - season_start).days
        
#         # Determine home and away teams
#         if game['Canucks_Home']:
#             home_team = game['Home_Team']
#             away_team = game['Away_Team']
#         else:
#             home_team = game['Home_Team']
#             away_team = game['Away_Team']  # 'VAN' should be one of these
        
#         # Create feature vector
#         features = create_feature_vector(home_team, away_team, team_stats, days_since_start)
        
#         # Predict the score
#         home_score, away_score, is_overtime = predict_score(features)
        
#         # Determine if Vancouver wins
#         vancouver_is_home = game['Canucks_Home']
#         vancouver_score = home_score if vancouver_is_home else away_score
#         opponent_score = away_score if vancouver_is_home else home_score
#         vancouver_win = vancouver_score > opponent_score
        
#         # Add to results
#         results.append({
#             'Date': game['Date'],
#             'Home_Team': home_team,
#             'Away_Team': away_team,
#             'Home_Score': home_score,
#             'Away_Score': away_score,
#             'Vancouver_Score': vancouver_score,
#             'Opponent_Score': opponent_score,
#             'Vancouver_Win': vancouver_win,
#             'Is_Overtime': is_overtime
#         })
    
#     # Convert to DataFrame
#     results_df = pd.DataFrame(results)
    
#     # Add win probability
#     results_df['Win_Probability'] = results_df.apply(
#         lambda row: calculate_win_probability(row['Home_Score'], row['Away_Score'], row['Home_Team'] == 'VAN' or row['Home_Team'] == 'Vancouver Canucks'),
#         axis=1
#     )
    
#     # Format results with OT indicator
#     results_df['Predicted_Score'] = results_df.apply(
#         lambda row: f"{row['Home_Team']} {row['Home_Score']} - {row['Away_Score']} {row['Away_Team']}{' (OT)' if row['Is_Overtime'] else ''}",
#         axis=1
#     )
    
#     return results_df

# # Function to calculate win probability
# def calculate_win_probability(home_score, away_score, canucks_home):
#     """
#     Calculate win probability based on predicted scores
#     """
#     diff = abs(home_score - away_score)
    
#     if (home_score > away_score and canucks_home) or (away_score > home_score and not canucks_home):
#         # Canucks win
#         if diff == 1:
#             # Close game
#             return min(0.5 + 0.15, 0.95)
#         else:
#             # Larger margin
#             return min(0.5 + 0.1 * diff, 0.95)
#     else:
#         # Canucks lose
#         if diff == 1:
#             # Close game
#             return max(0.5 - 0.15, 0.05)
#         else:
#             # Larger margin
#             return max(0.5 - 0.1 * diff, 0.05)

# # Run predictions and save to file
# def main(schedule_file):
#     """
#     Main function to run predictions and save results
#     """
#     try:
#         # Run predictions
#         results = predict_canucks_games(schedule_file)
        
#         # Calculate overall statistics
#         total_games = len(results)
        
#         if total_games == 0:
#             print("No future games found in the schedule.")
#             return pd.DataFrame()
            
#         wins = results['Vancouver_Win'].sum()
#         losses = total_games - wins
#         win_percentage = wins / total_games
        
#         ot_games = results['Is_Overtime'].sum()
#         ot_percentage = (ot_games / total_games) * 100
        
#         regulation_wins = sum((results['Vancouver_Win']) & (~results['Is_Overtime']))
#         ot_wins = sum((results['Vancouver_Win']) & (results['Is_Overtime']))
        
#         avg_goals_for = results['Vancouver_Score'].mean()
#         avg_goals_against = results['Opponent_Score'].mean()
        
#         # Print summary
#         print(f"\nVancouver Canucks Prediction Summary")
#         print(f"==================================")
#         print(f"Predicted Record: {wins:.0f}-{losses:.0f}")
#         print(f"Regulation Wins: {regulation_wins:.0f}")
#         print(f"Overtime/Shootout Wins: {ot_wins:.0f}")
#         print(f"Win Percentage: {win_percentage:.3f}")
#         print(f"Games Going to Overtime: {ot_games:.0f} ({ot_percentage:.1f}%)")
#         print(f"Average Goals For: {avg_goals_for:.2f}")
#         print(f"Average Goals Against: {avg_goals_against:.2f}")
#         print(f"==================================\n")
        
#         # Save results to Excel
#         output_file = 'vancouver_canucks_predictions.xlsx'
#         results.to_excel(output_file, index=False)
#         print(f"Detailed predictions saved to {output_file}")
        
#         return results
    
#     except Exception as e:
#         print(f"Error in main function: {e}")
#         import traceback
#         traceback.print_exc()
#         return pd.DataFrame()

# if __name__ == "__main__":
#     # Replace with your actual file path
#     schedule_file = "future_games.xlsx"
    
#     if os.path.exists(schedule_file):
#         predictions = main(schedule_file)
        
#         if not predictions.empty:
#             # Display the next 5 games
#             print("\nUpcoming 10 Games Predictions:")
#             for idx, game in predictions.head(10).iterrows():
#                 ot_indicator = " (OT)" if game['Is_Overtime'] else ""
#                 print(f"{game['Date'].strftime('%Y-%m-%d')}: {game['Home_Team']} {game['Home_Score']} - {game['Away_Score']} {game['Away_Team']}{ot_indicator} (Win Prob: {game['Win_Probability']:.2f})")
#     else:
#         print(f"Error: Schedule file '{schedule_file}' not found.")
#         print("Please provide a valid Excel file with Canucks schedule.")
#         print("Format required: Column 1 = Date, Column 2 = Home Team, Column 3 = Away Team")
